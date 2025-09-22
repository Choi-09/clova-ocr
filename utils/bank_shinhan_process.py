import json
import os
import re
import time
import uuid

import requests
from PyPDF2 import PdfReader, PdfWriter
from dotenv import load_dotenv
from openpyxl import Workbook

load_dotenv()

# 기본 설정
API_URL = os.getenv("API_URL")
SECRET_KEY = os.getenv("SECRET_KEY")


# PDF -> 단일 페이지 PDF로 분할
def split_pdf_pages(pdf_path, output_dir):
    """
    PDF 파일을 페이지별로 분할하여 개별 PDF 파일로 저장
    """
    os.makedirs(output_dir, exist_ok=True)
    reader = PdfReader(pdf_path)
    paths = []

    base_filename = os.path.splitext(os.path.basename(pdf_path))[0]

    for i, page in enumerate(reader.pages):
        path = os.path.join(output_dir, f"{base_filename}-{i + 1}.pdf")
        with open(path, "wb") as f:
            writer = PdfWriter()
            writer.add_page(page)
            writer.write(f)
        paths.append(path)
    return paths


# OCR 요청
def call_ocr_api(pdf_path):
    """
    네이버 CLOVA OCR API를 호출하여 PDF 파일에서 텍스트 추출
    """
    try:
        with open(pdf_path, 'rb') as file:
            payload = {
                'message': json.dumps({
                    'images': [{'format': 'pdf', 'name': 'demo'}],
                    'requestId': str(uuid.uuid4()),
                    'version': 'V2',
                    'timestamp': int(time.time() * 1000)
                }).encode('UTF-8')
            }
            files = [('file', file)]
            headers = {'X-OCR-SECRET': SECRET_KEY}
            response = requests.post(API_URL, headers=headers, data=payload, files=files)
            response.raise_for_status()  # HTTP 오류 발생 시 예외 발생
            # print(response.json())
            return response.json()
    except requests.exceptions.RequestException as e:
        print(f"[ERROR] OCR API 호출 실패: {e}")
        return None
    except Exception as e:
        print(f"[ERROR] OCR API 처리 중 예상치 못한 오류 발생: {e}")
        return None


# 기본 정보 추출
def extract_basic_info(text):
    patterns = {
        '계좌번호': r'계좌번호[:\s]*([\d\-]+)',
        '예금주': r'예금주 성명[:\s]*([^\s]+)',
        '상품명': r'상품명[:\s]*([^\s]+)',
        '조회기간': r'조회기간[:\s]*([\d\-]+)\s* \s*([\d\-]+)',
        '업무구분': r'업무구분 1:[:\s]*([\d:가-힣]+)'
    }
    info = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, text)
        if key == '조회기간' and match:
            info[key] = f"{match.group(1)} ~ {match.group(2)}"
        else:
            info[key] = match.group(1) if match else ''
    return info


# 거래내역 추출
# def extract_transaction_table_by_position(fields, y_threshold=10):
#     """
#     OCR fields를 이용해 거래내역 테이블을 헤더 기준으로 row별로 분리.
#     각 row는 merged_header 기준으로 셀 단위로 분리됨.
#     """
#     # 1. 헤더 subFields와 메인 텍스트 블록 추출
#     header_fields_from_subfields = []
#     main_table_text_block_content = ""
#
#     for field_item in fields:
#         if 'subFields' in field_item and field_item['subFields']:
#             for sf in field_item['subFields']:
#                 if sf.get('inferText') and sf.get('boundingPoly'):
#                     header_fields_from_subfields.append({
#                         'text': sf['inferText'].strip(),
#                         'x': sum(v.get('x', 0) for v in sf['boundingPoly']['vertices']) / 4,
#                         'y': sum(v.get('y', 0) for v in sf['boundingPoly']['vertices']) / 4
#                     })
#         # 긴 텍스트 블록은 메인 테이블로 간주
#         if '\n' in field_item.get('inferText', '') and len(field_item.get('inferText', '').split('\n')) > 3:
#             main_table_text_block_content = field_item['inferText']
#         # print("main_table_text_block_content: ", main_table_text_block_content)
#         # print("header_fields_from_subfields: ", header_fields_from_subfields)
#     if not header_fields_from_subfields or not main_table_text_block_content:
#         raise ValueError("헤더 또는 테이블 텍스트 블록을 찾을 수 없습니다.")
#
#     # 2. 헤더 행 찾기 (y 기준 그룹화)
#     items = sorted(header_fields_from_subfields, key=lambda x: x['y'])
#     rows, row, prev_y = [], [], None
#     for item in items:
#         if prev_y is None or abs(item['y'] - prev_y) <= y_threshold:
#             row.append(item)
#         else:
#             rows.append(row)
#             row = [item]
#         prev_y = item['y']
#     if row:
#         rows.append(row)
#
#     # 3. 헤더 행 선택
#     for row in rows:
#         text_line = ''.join([i['text'] for i in row])
#         if all(k in text_line for k in ['거래일자', '내용', '찾으신금액', '맡기신금액']):
#             header_row = sorted(row, key=lambda x: x['x'])
#             break
#     else:
#         raise ValueError("헤더를 찾을 수 없습니다.")
#
#     # 4. 헤더 병합 처리
#     merged_header, merged_x = [], []
#     i = 0
#     while i < len(header_row):
#         curr = header_row[i]['text']
#         if i + 1 < len(header_row):
#             next_text = header_row[i + 1]['text']
#             if curr + next_text in ['비고', '잔액']:
#                 merged_header.append(curr + next_text)
#                 merged_x.append((header_row[i]['x'] + header_row[i + 1]['x']) / 2)
#                 i += 2
#                 continue
#         merged_header.append(curr)
#         merged_x.append(header_row[i]['x'])
#         i += 1
#
#     # 5. 컬럼 경계 계산 (x 기준)
#     column_boundaries = []
#     for idx, x in enumerate(merged_x):
#         left = (merged_x[idx - 1] + x) / 2 if idx > 0 else -float('inf')
#         right = (x + merged_x[idx + 1]) / 2 if idx < len(merged_x) - 1 else float('inf')
#         column_boundaries.append((left, right))
#
#     # 6. 숫자 뒤 문자 분리 함수
#     def split_number_and_text(part):
#         match = re.match(r'₩([\d,]+)\s*(.*)', part)
#         if match:
#             number = match.group(1)
#             text = match.group(2).strip()
#             if text:
#                 return [number, text]
#             else:
#                 return [number]
#         else:
#             return [part]
#
#     # 7. 거래내역 행 그룹화
#     aligned_rows = []
#     # main_table_text_block_content를 subfield 단위로 word 추출
#     words = []
#     for field_item in fields:
#         if 'subFields' not in field_item:
#             continue
#         for sf in field_item['subFields']:
#             if sf.get('inferText') and sf.get('boundingPoly'):
#                 cx = sum(v.get('x', 0) for v in sf['boundingPoly']['vertices']) / 4
#                 cy = sum(v.get('y', 0) for v in sf['boundingPoly']['vertices']) / 4
#                 words.append({'text': sf['inferText'], 'x': cx, 'y': cy})
#
#     # y 기준 row 그룹화
#     words = sorted(words, key=lambda w: w['y'])
#     rows_grouped, current_row, prev_y = [], [], None
#     for w in words:
#         if prev_y is None or abs(w['y'] - prev_y) <= y_threshold:
#             current_row.append(w)
#         else:
#             rows_grouped.append(current_row)
#             current_row = [w]
#         prev_y = w['y']
#     if current_row:
#         rows_grouped.append(current_row)
#
#     # 8. 각 row에서 x 좌표 기준으로 컬럼에 맞춰 분배
#     for word_row in rows_grouped:
#         row_cells = [''] * len(merged_header)
#         for w in word_row:
#             for idx, (left, right) in enumerate(column_boundaries):
#                 if left <= w['x'] < right:
#                     split_parts = split_number_and_text(w['text'])
#                     for sp_idx, sp in enumerate(split_parts):
#                         target_idx = idx + sp_idx
#                         if target_idx < len(row_cells):
#                             if row_cells[target_idx]:
#                                 row_cells[target_idx] += ' ' + sp
#                             else:
#                                 row_cells[target_idx] = sp
#                     break
#         # 모든 컬럼이 빈 값이면 skip
#         if any(cell.strip() for cell in row_cells):
#             aligned_rows.append(row_cells)
#
#     return merged_header, aligned_rows
def extract_transaction_table_by_position(fields, y_threshold=5):
    header_fields = []
    main_text = ""

    # OCR 결과 필드 전부 처리
    for field in fields:
        if 'subFields' in field and field['subFields']:
            for sf in field['subFields']:
                if sf.get('inferText') and sf.get('boundingPoly'):
                    cx = sum(v.get('x', 0) for v in sf['boundingPoly']['vertices']) / 4
                    cy = sum(v.get('y', 0) for v in sf['boundingPoly']['vertices']) / 4
                    header_fields.append({'text': sf['inferText'].strip(), 'x': cx, 'y': cy})
        if 'inferText' in field and field['inferText']:
            main_text += field['inferText'] + " "

    if not header_fields:
        # subFields가 없으면 전체 텍스트 단어로 대체
        words = main_text.split()
        for idx, word in enumerate(words):
            header_fields.append({'text': word, 'x': idx * 10, 'y': 0})

    # y 기준으로 그룹화
    items = sorted(header_fields, key=lambda x: x['y'])
    rows, row, prev_y = [], [], None
    for item in items:
        if prev_y is None or abs(item['y'] - prev_y) <= y_threshold:
            row.append(item)
        else:
            rows.append(row)
            row = [item]
        prev_y = item['y']
    if row:
        rows.append(row)

    # 헤더 찾기 (키워드가 여러 row에 나눠져 있어도 병합)
    header_row = None
    header_keywords = ['거래일자', '내용', '찾으신금액', '맡기신금액']
    for r in rows:
        text_line = ''.join([i['text'] for i in r])
        if all(k in text_line for k in header_keywords):
            header_row = sorted(r, key=lambda x: x['x'])
            break
    if not header_row:
        # fallback: 첫 row를 헤더로
        header_row = sorted(rows[0], key=lambda x: x['x'])

    # 헤더와 x 좌표
    merged_header = [i['text'] for i in header_row]
    merged_x = [i['x'] for i in header_row]

    # 컬럼 경계 계산
    column_boundaries = []
    if len(merged_x) == 0:
        # 헤더 없음 → 빈 결과 반환
        return [], []
    elif len(merged_x) == 1:
        column_boundaries.append((-float('inf'), float('inf')))
    else:
        for idx, x in enumerate(merged_x):
            left = (merged_x[idx - 1] + x) / 2 if idx > 0 else -float('inf')
            right = (x + merged_x[idx + 1]) / 2 if idx < len(merged_x) - 1 else float('inf')
            column_boundaries.append((left, right))

    # 모든 word 추출
    words = []
    for field in fields:
        if 'subFields' not in field:
            continue
        for sf in field['subFields']:
            if sf.get('inferText') and sf.get('boundingPoly'):
                cx = sum(v.get('x', 0) for v in sf['boundingPoly']['vertices']) / 4
                cy = sum(v.get('y', 0) for v in sf['boundingPoly']['vertices']) / 4
                words.append({'text': sf['inferText'], 'x': cx, 'y': cy})

    # y 기준 row 그룹화
    words = sorted(words, key=lambda w: w['y'])
    rows_grouped, current_row, prev_y = [], [], None
    for w in words:
        if prev_y is None or abs(w['y'] - prev_y) <= y_threshold:
            current_row.append(w)
        else:
            rows_grouped.append(current_row)
            current_row = [w]
        prev_y = w['y']
    if current_row:
        rows_grouped.append(current_row)

    # 각 row를 column에 맞춰 분배
    aligned_rows = []
    for word_row in rows_grouped:
        row_cells = [''] * len(merged_header)
        for w in word_row:
            for idx, (left, right) in enumerate(column_boundaries):
                if left <= w['x'] < right:
                    if row_cells[idx]:
                        row_cells[idx] += ' ' + w['text']
                    else:
                        row_cells[idx] = w['text']
                    break
        if any(cell.strip() for cell in row_cells):
            aligned_rows.append(row_cells)

    return merged_header, aligned_rows


# 숫자 앞 문자 원화 기호로 바꾸기
def replace_W_before_number(text):
    """
    텍스트 내에서 숫자 앞에 있는 'W' 또는 '\' 문자를 '₩' (원화 기호)로 대체합니다.

    Args:
        text (str): 처리할 텍스트.

    Returns:
        str: 원화 기호로 대체된 텍스트.
    """
    return re.sub(r'[W\\](?=\d)', '₩', text)


# Excel 저장
def save_to_excel(wb, sheet_name, basic_info, header, rows):
    """
    추출된 기본 정보와 거래내역을 Excel 워크북에 시트로 저장합니다.

    Args:
        wb (openpyxl.Workbook): 저장할 Excel 워크북 객체.
        sheet_name (str): 생성할 시트의 이름.
        basic_info (dict): 기본 정보 딕셔너리.
        header (list): 테이블 헤더 리스트.
        rows (list): 테이블 데이터 행 리스트.
    """
    ws = wb.create_sheet(title=sheet_name)
    ws.append(["[기본 정보]"])
    for key, value in basic_info.items():
        ws.append([key, value])
    ws.append([])
    ws.append([f"[{sheet_name} - 거래 내역]"])
    ws.append(header)
    for row in rows[1:]:
        ws.append([replace_W_before_number(cell) for cell in row])


# 실행
def process_pdf_and_ocr(input_pdf_path, update_progress_callback=None):
    """
    PDF 파일을 처리하고 OCR을 수행하여 Excel 파일로 저장하는 메인 함수입니다.

    Args:
        input_pdf_path (str): 처리할 입력 PDF 파일의 경로.
        update_progress_callback (function, optional): 진행률을 업데이트하기 위한 콜백 함수.
                                                      Defaults to None.

    Returns:
        openpyxl.Workbook: 처리된 데이터가 포함된 Excel 워크북 객체.
    """
    temp_dir = os.path.join('temp_split', os.path.splitext(os.path.basename(input_pdf_path))[0])

    pdf_pages = split_pdf_pages(input_pdf_path, temp_dir)
    total_pages = len(pdf_pages)

    wb = Workbook()
    # 기본으로 생성된 시트 제거
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for idx, page_pdf in enumerate(pdf_pages):
        print(f"📄 처리 중: 페이지 {idx + 1}/{total_pages}")

        result = call_ocr_api(page_pdf)

        print("result: ", result)
        if not result or 'images' not in result:
            print(f"[⚠️] 페이지 {idx + 1} OCR 실패 또는 응답에 'images' 키 없음.")
            if update_progress_callback:
                update_progress_callback((idx + 1) / total_pages * 100)
            continue

        # 모든 필드를 하나의 리스트로 모읍니다.
        all_fields_from_ocr = []
        full_page_text_for_basic_info = ""

        for image_data in result['images']:
            if 'fields' in image_data:
                all_fields_from_ocr.extend(image_data['fields'])
                for field_item in image_data['fields']:
                    full_page_text_for_basic_info += field_item.get('inferText', '') + " "

            if 'title' in image_data and 'subFields' in image_data['title']:
                # title 필드 자체도 all_fields_from_ocr에 포함될 수 있습니다.
                # 여기서는 subFields만 추출하여 header_fields_from_subfields로 넘겨줄 필요가 없으므로,
                # 단순히 full_page_text_for_basic_info에 추가합니다.
                full_page_text_for_basic_info += image_data['title'].get('inferText', '') + " "

        # 페이지 처리 진행률 업데이트
        if update_progress_callback:
            progress = (idx + 1) / total_pages * 100
            update_progress_callback(int(progress))

        # 기본 정보 추출
        basic_info = extract_basic_info(full_page_text_for_basic_info)
        print(f'페이지 {idx + 1}에서 추출된 텍스트 (일부): {full_page_text_for_basic_info[:200]}...')  # 디버깅을 위해 처음 200자 인쇄

        try:
            # extract_transaction_table_by_position 함수에 모든 필드를 전달합니다.
            header, rows = extract_transaction_table_by_position(all_fields_from_ocr)
            print(f"✅ 페이지 {idx + 1} 거래 내역 추출 완료.")
            save_to_excel(wb, f"페이지 {idx + 1}", basic_info, header, rows)
            print(f"✅ 페이지 {idx + 1} 거래 내역 저장 완료.")
        except ValueError as ve:
            print(f"[❌] 페이지 {idx + 1} 거래 내역 처리 오류: {ve}")
        except Exception as e:
            print(f"[❌] 페이지 {idx + 1} 처리 중 예상치 못한 오류 발생: {e}")

    return wb
